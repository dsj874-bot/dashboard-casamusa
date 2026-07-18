"""
Genera/regenera data/NE_x_Facturar.xlsx a partir de los vendedores
oficiales definidos en data_loader.VEND_HOME (una fila por vendedor
"home" de cada sucursal + una fila "OTROS").

Uso: python generar_plantilla_ne.py
Se puede volver a correr cuando cambie el equipo de ventas; si el
archivo data/NE_x_Facturar.xlsx ya existe, los montos ya cargados
para vendedores que siguen existiendo se conservan.
"""
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Protection
from openpyxl.utils import get_column_letter

from data_loader import VEND_HOME, ORDEN_SUCURSALES

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DESTINO  = os.path.join(BASE_DIR, "data", "NE_x_Facturar.xlsx")


def montos_existentes():
    """Si ya existe un NE_x_Facturar.xlsx, rescata los montos actuales."""
    if not os.path.exists(DESTINO):
        return {}
    wb = openpyxl.load_workbook(DESTINO, data_only=True)
    ws = wb.active
    montos = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None:
            continue
        suc, vend, monto = row[0], row[1], row[2]
        montos[(str(suc).strip(), str(vend).strip())] = monto or 0
    return montos


def construir_filas():
    previos = montos_existentes()
    filas = []
    for suc in ORDEN_SUCURSALES:
        vendedores = sorted(VEND_HOME.get(suc, set()))
        for v in vendedores:
            monto = previos.get((suc, v), 0)
            filas.append((suc, v, monto))
        # Fila OTROS para negocios de vendedores no listados en esa sucursal
        monto_otros = previos.get((suc, "OTROS"), 0)
        filas.append((suc, "OTROS", monto_otros))
    return filas


def generar():
    filas = construir_filas()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NE x Facturar"

    encabezados = ["Sucursal", "Vendedor", "Monto NE"]
    ws.append(encabezados)

    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, 4):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center")

    fila_actual = 2
    suc_anterior = None
    fill_par = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
    for suc, vend, monto in filas:
        ws.cell(row=fila_actual, column=1, value=suc)
        ws.cell(row=fila_actual, column=2, value=vend)
        celda_monto = ws.cell(row=fila_actual, column=3, value=monto)
        celda_monto.number_format = "#,##0"

        # Sombrear por sucursal para que sea facil de leer
        if suc != suc_anterior:
            usar_sombra = (ORDEN_SUCURSALES.index(suc) % 2 == 1)
        if usar_sombra:
            for col in range(1, 4):
                ws.cell(row=fila_actual, column=col).fill = fill_par
        suc_anterior = suc

        # Bloquear Sucursal y Vendedor; dejar editable solo Monto NE
        ws.cell(row=fila_actual, column=1).protection = Protection(locked=True)
        ws.cell(row=fila_actual, column=2).protection = Protection(locked=True)
        celda_monto.protection = Protection(locked=False)

        fila_actual += 1

    # Encabezado tambien bloqueado (da lo mismo, la hoja completa parte bloqueada)
    for col in range(1, 4):
        ws.cell(row=1, column=col).protection = Protection(locked=True)

    ws.protection.sheet = True
    ws.protection.enable()  # sin contraseña: solo evita ediciones accidentales

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 30
    ws.column_dimensions["C"].width = 16

    ws.freeze_panes = "A2"

    # Nota de instrucciones en una segunda hoja
    ws2 = wb.create_sheet("Instrucciones")
    ws2["A1"] = "Instrucciones"
    ws2["A1"].font = Font(bold=True, size=13)
    notas = [
        "",
        "Esta planilla se actualiza 1 vez por semana con el monto de Negocios",
        "Ganados aun no facturados (NE x Facturar), por sucursal y vendedor.",
        "",
        "Solo debes escribir montos en la columna 'Monto NE' (hoja NE x Facturar).",
        "No agregues, borres ni reordenes filas, y no edites Sucursal ni Vendedor",
        "(esas columnas estan bloqueadas para evitar errores de tipeo).",
        "",
        "Si un vendedor nuevo entra al equipo o uno se va, avisa a Sistemas",
        "para regenerar la planilla con la lista de vendedores actualizada.",
        "",
        "Guarda el archivo con el mismo nombre (NE_x_Facturar.xlsx) en la",
        "misma carpeta, reemplazando el anterior.",
    ]
    for i, linea in enumerate(notas, start=1):
        ws2.cell(row=i, column=1, value=linea)
    ws2.column_dimensions["A"].width = 75

    os.makedirs(os.path.dirname(DESTINO), exist_ok=True)
    wb.save(DESTINO)
    print(f"Generado: {DESTINO}")
    print(f"Filas: {len(filas)}")


if __name__ == "__main__":
    generar()
